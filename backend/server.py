from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, date
from decimal import Decimal
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Define Models
class Transaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tipo: str  # "receita" or "despesa"
    valor: float
    descricao: str
    data: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    mes: int
    ano: int

class TransactionCreate(BaseModel):
    tipo: str
    valor: float
    descricao: str
    data: Optional[datetime] = None

class FixedExpense(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    valor: float
    data_vencimento: int  # dia do mês (1-31)
    pago: bool = False
    mes: int
    ano: int
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FixedExpenseCreate(BaseModel):
    nome: str
    valor: float
    data_vencimento: int
    mes: int
    ano: int

class FixedExpenseUpdate(BaseModel):
    pago: bool

class AlertConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    limite_mensal: float
    mes: int
    ano: int
    ativo: bool = True

class AlertConfigCreate(BaseModel):
    limite_mensal: float
    mes: int
    ano: int

class MonthlyReport(BaseModel):
    mes: int
    ano: int
    total_receitas: float
    total_despesas: float
    saldo: float
    transacoes: List[Transaction]
    despesas_fixas: List[FixedExpense]
    total_despesas_fixas: float
    despesas_fixas_pagas: float
    despesas_fixas_pendentes: float
    limite_excedido: bool = False
    limite_configurado: Optional[float] = None

# Helper functions
def prepare_for_mongo(data):
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, datetime):
                data[key] = value.isoformat()
    return data

def parse_from_mongo(item):
    if isinstance(item.get('data'), str):
        item['data'] = datetime.fromisoformat(item['data'])
    if isinstance(item.get('created_at'), str):
        item['created_at'] = datetime.fromisoformat(item['created_at'])
    return item

def format_currency_br(value):
    """Format currency in Brazilian Real format"""
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# Transaction endpoints
@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(transaction: TransactionCreate):
    transaction_dict = transaction.dict()
    
    # Set date if not provided
    if not transaction_dict.get('data'):
        transaction_dict['data'] = datetime.now(timezone.utc)
    
    # Extract month and year
    data = transaction_dict['data']
    transaction_dict['mes'] = data.month
    transaction_dict['ano'] = data.year
    
    # Create transaction object
    trans_obj = Transaction(**transaction_dict)
    
    # Prepare for MongoDB
    trans_dict = prepare_for_mongo(trans_obj.dict())
    
    # Insert into database
    await db.transactions.insert_one(trans_dict)
    
    return trans_obj

@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(mes: Optional[int] = None, ano: Optional[int] = None):
    query = {}
    if mes and ano:
        query = {"mes": mes, "ano": ano}
    elif ano:
        query = {"ano": ano}
    
    transactions = await db.transactions.find(query).sort("data", -1).to_list(1000)
    
    # Parse from MongoDB
    parsed_transactions = [parse_from_mongo(trans) for trans in transactions]
    
    return [Transaction(**trans) for trans in parsed_transactions]

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str):
    result = await db.transactions.delete_one({"id": transaction_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"message": "Transaction deleted successfully"}

# Fixed Expenses endpoints
@api_router.post("/fixed-expenses", response_model=FixedExpense)
async def create_fixed_expense(expense: FixedExpenseCreate):
    expense_dict = expense.dict()
    expense_obj = FixedExpense(**expense_dict)
    
    # Prepare for MongoDB
    expense_dict = prepare_for_mongo(expense_obj.dict())
    
    # Insert into database
    await db.fixed_expenses.insert_one(expense_dict)
    
    return expense_obj

@api_router.get("/fixed-expenses", response_model=List[FixedExpense])
async def get_fixed_expenses(mes: Optional[int] = None, ano: Optional[int] = None):
    query = {}
    if mes and ano:
        query = {"mes": mes, "ano": ano}
    elif ano:
        query = {"ano": ano}
    
    expenses = await db.fixed_expenses.find(query).sort("data_vencimento", 1).to_list(1000)
    
    # Parse from MongoDB
    parsed_expenses = [parse_from_mongo(expense) for expense in expenses]
    
    return [FixedExpense(**expense) for expense in parsed_expenses]

@api_router.put("/fixed-expenses/{expense_id}", response_model=FixedExpense)
async def update_fixed_expense(expense_id: str, update: FixedExpenseUpdate):
    result = await db.fixed_expenses.update_one(
        {"id": expense_id},
        {"$set": {"pago": update.pago}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fixed expense not found")
    
    # Get updated expense
    expense = await db.fixed_expenses.find_one({"id": expense_id})
    if expense:
        parsed_expense = parse_from_mongo(expense)
        return FixedExpense(**parsed_expense)
    
    raise HTTPException(status_code=404, detail="Fixed expense not found")

@api_router.delete("/fixed-expenses/{expense_id}")
async def delete_fixed_expense(expense_id: str):
    result = await db.fixed_expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fixed expense not found")
    return {"message": "Fixed expense deleted successfully"}

# Alert configuration endpoints
@api_router.post("/alerts", response_model=AlertConfig)
async def create_alert_config(alert: AlertConfigCreate):
    alert_dict = alert.dict()
    alert_obj = AlertConfig(**alert_dict)
    
    # Remove any existing alert for this month/year
    await db.alerts.delete_many({"mes": alert.mes, "ano": alert.ano})
    
    # Insert new alert
    await db.alerts.insert_one(alert_obj.dict())
    
    return alert_obj

@api_router.get("/alerts", response_model=List[AlertConfig])
async def get_alert_configs():
    alerts = await db.alerts.find().to_list(1000)
    return [AlertConfig(**alert) for alert in alerts]

@api_router.get("/alerts/{mes}/{ano}", response_model=Optional[AlertConfig])
async def get_alert_config(mes: int, ano: int):
    alert = await db.alerts.find_one({"mes": mes, "ano": ano})
    if alert:
        return AlertConfig(**alert)
    return None

# Reports endpoint
@api_router.get("/reports/{mes}/{ano}", response_model=MonthlyReport)
async def get_monthly_report(mes: int, ano: int):
    # Get transactions for the month
    transactions = await db.transactions.find({"mes": mes, "ano": ano}).to_list(1000)
    parsed_transactions = [parse_from_mongo(trans) for trans in transactions]
    trans_objects = [Transaction(**trans) for trans in parsed_transactions]
    
    # Get fixed expenses for the month
    fixed_expenses = await db.fixed_expenses.find({"mes": mes, "ano": ano}).to_list(1000)
    parsed_fixed_expenses = [parse_from_mongo(expense) for expense in fixed_expenses]
    fixed_expense_objects = [FixedExpense(**expense) for expense in parsed_fixed_expenses]
    
    # Calculate totals
    total_receitas = sum(t.valor for t in trans_objects if t.tipo == "receita")
    total_despesas = sum(t.valor for t in trans_objects if t.tipo == "despesa")
    
    # Calculate fixed expenses totals
    total_despesas_fixas = sum(e.valor for e in fixed_expense_objects)
    despesas_fixas_pagas = sum(e.valor for e in fixed_expense_objects if e.pago)
    despesas_fixas_pendentes = sum(e.valor for e in fixed_expense_objects if not e.pago)
    
    # Total including fixed expenses
    total_despesas_all = total_despesas + total_despesas_fixas
    saldo = total_receitas - total_despesas_all
    
    # Check alert configuration
    alert_config = await db.alerts.find_one({"mes": mes, "ano": ano, "ativo": True})
    limite_excedido = False
    limite_configurado = None
    
    if alert_config:
        limite_configurado = alert_config["limite_mensal"]
        limite_excedido = total_despesas_all > limite_configurado
    
    return MonthlyReport(
        mes=mes,
        ano=ano,
        total_receitas=total_receitas,
        total_despesas=total_despesas,
        saldo=saldo,
        transacoes=trans_objects,
        despesas_fixas=fixed_expense_objects,
        total_despesas_fixas=total_despesas_fixas,
        despesas_fixas_pagas=despesas_fixas_pagas,
        despesas_fixas_pendentes=despesas_fixas_pendentes,
        limite_excedido=limite_excedido,
        limite_configurado=limite_configurado
    )

# Dashboard data endpoint
@api_router.get("/dashboard/{ano}")
async def get_dashboard_data(ano: int):
    monthly_data = []
    
    for mes in range(1, 13):
        report = await get_monthly_report(mes, ano)
        monthly_data.append({
            "mes": mes,
            "receitas": report.total_receitas,
            "despesas": report.total_despesas + report.total_despesas_fixas,
            "despesas_variaveis": report.total_despesas,
            "despesas_fixas": report.total_despesas_fixas,
            "saldo": report.saldo
        })
    
    return {"ano": ano, "dados_mensais": monthly_data}

# Export endpoints
@api_router.get("/export/csv/{mes}/{ano}")
async def export_csv(mes: int, ano: int):
    """Export monthly data to CSV format"""
    report = await get_monthly_report(mes, ano)
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')  # Using semicolon for better Excel compatibility
    
    # Write headers and summary
    writer.writerow([f"RELATÓRIO FINANCEIRO - {['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes]} {ano}"])
    writer.writerow([])
    
    # Summary section
    writer.writerow(["RESUMO MENSAL"])
    writer.writerow(["Receitas Totais", f"R$ {report.total_receitas:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    writer.writerow(["Despesas Variáveis", f"R$ {report.total_despesas:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    writer.writerow(["Despesas Fixas", f"R$ {report.total_despesas_fixas:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    writer.writerow(["Saldo Final", f"R$ {report.saldo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    writer.writerow([])
    
    # Transactions section
    writer.writerow(["TRANSAÇÕES"])
    writer.writerow(["Data", "Tipo", "Descrição", "Valor"])
    
    for transaction in report.transacoes:
        formatted_date = transaction.data.strftime("%d/%m/%Y")
        formatted_value = f"R$ {transaction.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        if transaction.tipo == "despesa":
            formatted_value = f"-{formatted_value}"
        writer.writerow([formatted_date, transaction.tipo.title(), transaction.descricao, formatted_value])
    
    writer.writerow([])
    
    # Fixed expenses section
    writer.writerow(["DESPESAS FIXAS"])
    writer.writerow(["Nome", "Vencimento", "Valor", "Status"])
    
    for expense in report.despesas_fixas:
        formatted_value = f"R$ {expense.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        status = "Pago" if expense.pago else "Pendente"
        writer.writerow([expense.nome, f"Dia {expense.data_vencimento}", formatted_value, status])
    
    # Prepare response
    output.seek(0)
    month_name = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes]
    filename = f"financas_{month_name}_{ano}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/excel/{mes}/{ano}")
async def export_excel(mes: int, ano: int):
    """Export monthly data to Excel format"""
    report = await get_monthly_report(mes, ano)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    month_names = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    month_name = month_names[mes]
    ws.title = f"{month_name} {ano}"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_font = Font(bold=True, size=12)
    currency_font = Font(size=11)
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:E{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = f"RELATÓRIO FINANCEIRO - {month_name} {ano}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    row += 3
    
    # Summary section
    ws[f'A{row}'] = "RESUMO MENSAL"
    ws[f'A{row}'].font = summary_font
    row += 1
    
    summary_data = [
        ("Receitas Totais", report.total_receitas, "4CAF50"),
        ("Despesas Variáveis", report.total_despesas, "F44336"),
        ("Despesas Fixas", report.total_despesas_fixas, "FF9800"),
        ("Saldo Final", report.saldo, "4CAF50" if report.saldo >= 0 else "F44336")
    ]
    
    for label, value, color in summary_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, color=color)
        row += 1
    
    row += 2
    
    # Transactions section
    ws[f'A{row}'] = "TRANSAÇÕES"
    ws[f'A{row}'].font = summary_font
    row += 1
    
    # Transaction headers
    headers = ["Data", "Tipo", "Descrição", "Valor"]
    for col, header in enumerate(headers, 1):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Transaction data
    for transaction in report.transacoes:
        ws[f'A{row}'] = transaction.data.strftime("%d/%m/%Y")
        ws[f'B{row}'] = transaction.tipo.title()
        ws[f'C{row}'] = transaction.descricao
        ws[f'D{row}'] = transaction.valor if transaction.tipo == "receita" else -transaction.valor
        ws[f'D{row}'].number_format = 'R$ #,##0.00'
        
        # Color coding
        if transaction.tipo == "receita":
            ws[f'D{row}'].font = Font(color="4CAF50")
        else:
            ws[f'D{row}'].font = Font(color="F44336")
        
        row += 1
    
    row += 2
    
    # Fixed expenses section
    ws[f'A{row}'] = "DESPESAS FIXAS"
    ws[f'A{row}'].font = summary_font
    row += 1
    
    # Fixed expenses headers
    headers = ["Nome", "Vencimento", "Valor", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws[f'{get_column_letter(col)}{row}']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    
    # Fixed expenses data
    for expense in report.despesas_fixas:
        ws[f'A{row}'] = expense.nome
        ws[f'B{row}'] = f"Dia {expense.data_vencimento}"
        ws[f'C{row}'] = expense.valor
        ws[f'C{row}'].number_format = 'R$ #,##0.00'
        ws[f'D{row}'] = "Pago" if expense.pago else "Pendente"
        
        # Color coding for status
        if expense.pago:
            ws[f'D{row}'].font = Font(color="4CAF50")
        else:
            ws[f'D{row}'].font = Font(color="FF9800")
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"financas_{month_name}_{ano}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()